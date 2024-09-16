import pandas as pd
from Employee import Employee
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import os

employees = dict()



file_path = input("Enter File Path :") 
first_day = input("Enter First Day in formate 'm/d/y':")
last_day = input("Enter Last Day in formate 'm/d/y':")

# dayes : ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
exceptions_dayes=['Friday', 'Saturday']
exceptions_dates=[]

#choice = input("Enter more details? [Y/N] :")
#if choice.lower() == "y":
#    print("1- Enter exceptions_days.")
#    print("2- Enter exceptions_dates.")
#    while True:
#        choice = input("Enter your choice [1 or 2] :")
#        if choice in ['1','2']:
#            break
#    if choice == '1':
        



df = pd.read_excel(file_path, sheet_name = 0, engine='xlrd')
    
count = 0

for index, row in df.iterrows():
    row = row.to_dict()
    id = row['No.']
    if id not in employees:
        employees[id] = Employee(id=id, name=row['Name'],department=row['Department'])
    employees[id].add_log(date_str=row['Date/Time'])

for key in sorted(employees):
    employees[key].add_empty_days(first_day,last_day)


wb = Workbook()
ws = wb.active

column_names = ["No.", "Name", "Department", "date", "check_in", "check_out", "worked_hours", "late"]
ws.append(column_names)

fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
font = Font(color="FFFFFF")

columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
for i in range(0,len(columns)):
    ws.column_dimensions[columns[i]].width = 25


for key in sorted(employees):
    emp_data = employees[key].fetch_data(exceptions_dayes = exceptions_dayes,exceptions_dates = exceptions_dates)
    total_late = 0
    for row in emp_data:
        ws.append(row)
        late_time = row[-1].split(':')
        if len(late_time) > 1:
            total_late += int(late_time[0]) * 60 + int(late_time[1])
    
    ws.append(['','','','','','','',f"{total_late // 60}:{total_late % 60}"]) # empty row
    for cell in ws[ws.max_row]:
        cell.fill = fill
        cell.font = font


save_path = "output"
i = 1
while True:
    if not os.path.isfile(save_path + str(i) + '.xlsx'):
        break
    i += 1

wb.save(save_path + str(i) + '.xlsx')


